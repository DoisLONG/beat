# Copyright (C) 2026 Intel Corporation
# SPDX-License-Identifier: Apache-2.0

import asyncio
import os
import re
import tempfile
import shutil
import time
import traceback
from abc import ABC
from pathlib import Path
import camelot
import math

import pandas as pd
import requests
from bs4 import BeautifulSoup
from fastapi import HTTPException

from comps import CustomLogger
from comps.dataprep import magic_pdf_parse_util
from comps.dataprep.config import (
    MAGIC_PDF_PARSE_CONNECT_TIMEOUT,
    MAGIC_PDF_PARSE_LANG,
    MAGIC_PDF_PARSE_METHOD,
    MAGIC_PDF_PARSE_READ_TIMEOUT,
    MAGIC_PDF_PARSE_RETRIES,
    MAGIC_PDF_PARSE_RETRY_BACKOFF_SECONDS,
    MAGIC_PDF_PARSE_URL,
    MAGIC_PDF_REMOTE_ENABLED,
    get_dataprep_llm_config,
    get_llm_extra_body,
    get_client,
)
from comps.dataprep.loaders.header_text_splitter import DocHierarchyParser, MarkdownHierarchyParser, \
    WordTableExtractorLite, MagicPdfTableExtractor
from openpyxl import load_workbook

from comps.dataprep.prompt.prompt_manager import PromptRegistry, PromptKey, Lang
from comps.dataprep.utils import normalize, _load_doc, _load_docx, \
     detect_language_from_preview, try_parse_json
from comps.oss_manager import oss_manager

logger = CustomLogger("dataprep-loader", os.getenv("LOG_LEVEL", "INFO"))


def get_excel_engine(file_path: str) -> str:
    file_ext = Path(file_path).suffix.lower()
    if file_ext == ".xlsx":
        return "openpyxl"
    if file_ext == ".xls":
        return "xlrd"
    raise ValueError(f"不支持的文件格式: {file_ext},仅支持 .xlsx 和 .xls")


def is_html_disguised_excel(file_path: str, sniff_size: int = 4096) -> bool:
    try:
        with open(file_path, "rb") as file_obj:
            head = file_obj.read(sniff_size)
    except OSError:
        return False

    normalized = head.lstrip().lower()
    return normalized.startswith((b"<html", b"<!doctype html", b"<table")) or b"<html" in normalized or b"<table" in normalized


def read_html_table_as_dataframe(
    file_path: str,
    *,
    table_index: int = 0,
    skiprows: int = 0,
    nrows: int | None = None,
) -> pd.DataFrame:
    with open(file_path, "r", encoding="utf-8", errors="ignore") as file_obj:
        html_content = file_obj.read()

    soup = BeautifulSoup(html_content, "html.parser")
    tables = soup.find_all("table")
    if not tables:
        raise ValueError(f"HTML伪装的Excel文件中未找到表格: {file_path}")
    if table_index >= len(tables):
        raise ValueError(f"HTML伪装的Excel文件表格索引越界: {table_index}, 总表格数: {len(tables)}")

    rows: list[list[str]] = []
    for tr in tables[table_index].find_all("tr"):
        cells = [cell.get_text(strip=True) for cell in tr.find_all(["td", "th"])]
        if cells:
            rows.append(cells)

    if skiprows:
        rows = rows[skiprows:]
    if nrows is not None:
        rows = rows[:nrows]

    return pd.DataFrame(rows, dtype=object)


def read_excel_like_dataframe(
    file_path: str,
    *,
    sheet_name: int | str = 0,
    header: int | list[int] | None = None,
    nrows: int | None = None,
    skiprows: int = 0,
) -> pd.DataFrame:
    if is_html_disguised_excel(file_path):
        if header is not None:
            raise ValueError("HTML伪装的Excel文件暂不支持显式header解析")
        if sheet_name not in (0, "Sheet1"):
            raise ValueError(f"HTML伪装的Excel文件仅支持首个表格读取，收到 sheet_name={sheet_name}")
        return read_html_table_as_dataframe(
            file_path,
            table_index=0,
            skiprows=skiprows,
            nrows=nrows,
        )

    return pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=header,
        nrows=nrows,
        skiprows=skiprows,
        engine=get_excel_engine(file_path),
    )

# 公共辅助函数: 从 OSS 下载文件写入临时目录
async def download_oss_to_temp(
    oss_uri: str,
    filename: str | None = None,
    small_file_limit: int = 10 * 1024 * 1024
) -> tuple[str, str]:
    """
    下载 OSS 文件到临时目录并返回 (tmp_dir, temp_file_path)
    - 自动判断文件大小
    - 小文件走 get_bytes
    - 大文件走流式下载
    """
    object_key = oss_manager.parse_key(oss_uri)
    if not object_key:
        raise FileNotFoundError(f"非法 OSS URI: {oss_uri}")

    # 1 获取文件大小（关键）
    size = await oss_manager.get_object_size(object_key)

    tmp_dir = tempfile.mkdtemp()
    try:
        if not filename:
            filename = os.path.basename(object_key)

        temp_path = os.path.join(tmp_dir, filename)

        # 2 小文件：内存方式
        if size <= small_file_limit:
            data = await oss_manager.get_bytes(oss_uri, max_size=small_file_limit)
            if not data:
                raise FileNotFoundError(f"OSS 文件不存在: {oss_uri}")

            with open(temp_path, "wb") as f:
                f.write(data)

        # 3 大文件：流式下载
        else:
            ok = await oss_manager.download_large_file(
                oss_uri,
                temp_path
            )
            if not ok:
                raise IOError(f"OSS 大文件下载失败: {oss_uri}")
        
        return tmp_dir, temp_path
    except Exception as e:
        cleanup_tmp_dir(tmp_dir)
        raise e

def cleanup_tmp_dir(tmp_dir: str):
    """安全清理临时目录及其内容。"""
    if tmp_dir and os.path.exists(tmp_dir):
        shutil.rmtree(tmp_dir, ignore_errors=True)

class BaseDataLoader(ABC):
    """
    Abstract base class for data loaders.
    """

    async def validate_headers(self,file_contents,**kwargs):
        """
        Abstract method to validate headers.
        """
        raise NotImplementedError("Subclasses must implement this method.")

    async def load_data(self, file_path: str,**kwargs):
        """
        Abstract method to load data.
        """
        raise NotImplementedError("Subclasses must implement this method.")

    async def preprocess_data(self,file_path:str,**kwargs):
        """
        Abstract method to preprocess data.
        """
        raise NotImplementedError("Subclasses must implement this method.")


class SopExcelDataUniversalLoader(BaseDataLoader):
    """
    通用版SOPExcel加载器
    """

    async def validate_headers(self, file_contents, **kwargs):
        """通过大模型，获取表头和数据行开始部分"""
        # 检查文件格式是否为xlsx或xls
        file_path = kwargs.get("file_path")
        client = kwargs.get("client")
        model_name = kwargs.get("model_name")
        heads, end_row, lang = await self._read_excel_with_merged_header(file_path, client, model_name)
        return heads, end_row, lang

    async def load_data(self, file_path: str, **kwargs):
        """
        Load SOP data from a local Excel file path.
        """
        db_client = kwargs.get("db_client")
        sop_id = kwargs.get("sop_id")
        client = kwargs.get("client")
        llm_config = get_dataprep_llm_config()
        model_name = kwargs.get("model_name") or llm_config.model

        if not client:
            client = get_client()

        try:
            heads, end_row, lang = await self.validate_headers(
                file_contents="",
                file_path=file_path,
                client=client,
                model_name=model_name,
            )
            if db_client and sop_id:
                db_client.update_percent_by_id(sop_id, "30%", lang=kwargs.get("lang", "zh"))
            structured_rows = await self.preprocess_data(file_path, heads=heads, end_row=end_row)
            structured_rows = await self._clear_dispimg_values_except_content(structured_rows)
            return structured_rows, lang
        except Exception as e:
            logger.error(f"Excel数据加载失败: {file_path}, 错误: {e}")
            raise

    async def preprocess_data(self, file_path: str, **kwargs):
        """
        Preprocess the loaded SOP Excel data.
        """
        heads = kwargs.get("heads")
        end_row = kwargs.get("end_row")

        try:
            for head in heads:
                if head == "nan":
                    continue
            df = read_excel_like_dataframe(file_path, skiprows=end_row, header=None)
            structured: list[dict[str, object]] = []
            prev_row_dict: dict[str, object] = {}

            for idx, row in df.iterrows():
                row_dict: dict[str, object] = {}

                for col_idx, head in enumerate(heads):
                    if col_idx < len(row):
                        value = row.iloc[col_idx]
                        if value is None or (isinstance(value, float) and math.isnan(value)) or str(value).strip().lower() == "nan" or str(value).strip() == "":
                            row_dict[head] = prev_row_dict.get(head, "")
                        else:
                            row_dict[head] = value
                    else:
                        row_dict[head] = prev_row_dict.get(head, "")

                row_dict["行号"] = idx + end_row + 1
                row_dict["内容"] = "；".join(
                    f"{key}:{value}" for key, value in row_dict.items() if key != "行号" and value not in ("", None)
                )

                structured.append(row_dict)
                prev_row_dict = row_dict

            return structured
        except Exception as e:
            logger.info(traceback.format_exc())
            raise HTTPException(status_code=500, detail=f"Excel解析失败: {e}")

    async def _clear_dispimg_values_except_content(self, data_list):
        """
        遍历字典列表，清空除 '内容' 外，值中包含 'DISPIMG' 的字段。
        """
        for item in data_list:
            for key, value in item.items():
                if key != "内容" and isinstance(value, str) and "DISPIMG" in value:
                    item[key] = ""
        return data_list

    async def _read_excel_with_merged_header(self, excel_path, client, model_name, sheet_name=0, max_rows=20):
        """
        读取 Excel，并识别表头行、数据起始行，同时处理合并单元格。
        """
        df_preview = read_excel_like_dataframe(
            excel_path,
            sheet_name=sheet_name,
            header=None,
            nrows=max_rows,
        )
        lang = await detect_language_from_preview(df_preview, client)
        logger.info(f"检测到的语言: {lang}")
        preview_text = df_preview.astype(str).values.tolist()
        prompt = PromptRegistry.get(PromptKey.GET_EXCEL_HEAD_AND_CONTENT_235B, Lang(lang), preview_text=preview_text)
        response = client.chat.completions.create(
            model=model_name,
            messages=[{"role": "user", "content": prompt}],
            extra_body=get_llm_extra_body(model_name),
            response_format={"type": "json_object"},
            temperature=0,
        )

        answer_text = response.choices[0].message.content.strip()
        result = try_parse_json(answer_text)
        return result.get("heads", []), result.get("end_row", 0), lang


class SOPPdfDataLoader(SopExcelDataUniversalLoader):
    """
    Data loader for SOP PDF documents.
    """

    async def load_data(self, file_path: str, **kwargs):
        """
        Load SOP data from a local PDF file path.
        """
        db_client = kwargs.get("db_client")
        sop_id = kwargs.get("sop_id")
        client = kwargs.get("client")
        llm_config = get_dataprep_llm_config()
        model_name = kwargs.get("model_name") or llm_config.model

        if not client:
            client = get_client()

        tmp_dir = tempfile.mkdtemp()
        try:
            tables = camelot.read_pdf(file_path, pages="all", flavor="lattice")
            if len(tables) == 0:
                tables = camelot.read_pdf(file_path, pages="all", flavor="stream")
            if len(tables) == 0:
                raise ValueError(f"PDF文件中未找到表格: {file_path}")

            tables = self._merge_single_table_with_header(tables)
            excel_filename = Path(file_path).stem + "_extracted.xlsx"
            temp_excel_path = os.path.join(tmp_dir, excel_filename)

            with pd.ExcelWriter(temp_excel_path, engine="openpyxl") as writer:
                for index, table in enumerate(tables):
                    dataframe = table.df.dropna(how="all").dropna(axis=1, how="all")
                    sheet_name = f"Table_{index + 1}" if len(tables) > 1 else "Sheet1"
                    dataframe.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

            heads, end_row, lang = await self.validate_headers(
                file_contents="",
                file_path=temp_excel_path,
                client=client,
                model_name=model_name,
            )
            if db_client and sop_id:
                db_client.update_percent_by_id(sop_id, "30%", lang=kwargs.get("lang", "zh"))

            structured_rows = await self.preprocess_data(temp_excel_path, heads=heads, end_row=end_row)
            structured_rows = await self._clear_dispimg_values_except_content(structured_rows)
            return structured_rows, lang
        except Exception as e:
            logger.error(f"PDF数据加载失败: {file_path}, 错误: {e}")
            raise
        finally:
            cleanup_tmp_dir(tmp_dir)

    def _merge_single_table_with_header(self, tables):
        """
        PDF 中只有一张表，但跨多页。
        第 1 页包含表头，后续页不包含表头。
        合并为一个 Camelot Table 对象。
        """
        base_df = tables[0].df.dropna(how="all").dropna(axis=1, how="all")
        for table in tables[1:]:
            dataframe = table.df.dropna(how="all").dropna(axis=1, how="all")
            if len(dataframe) > 1:
                dataframe = dataframe.iloc[1:]
            base_df = pd.concat([base_df, dataframe], ignore_index=True)

        merged_table = tables[0]
        merged_table.df = base_df
        return [merged_table]


class RISKExcelDataLoader(BaseDataLoader):
    """
    Data loader for RISK Excel documents.
    """

    def validate_headers(self,file_contents,**kwargs):

        target_headers = []
        file_path = kwargs.get("file_path")
        wb = load_workbook(file_path)
        ws = wb.active

        # 要检查的列
        cols = ['A','B', 'C', 'D', 'E','F','G']  #
        expected = ['序号', '标准作业卡名称', '事故案例收集', '风险因素描述', '风险控制措施','风险控制措施','是否落实']

        errors = []
        # 第一部分：A列到E列，从第2行开始
        row = 2
        previous_header = ""
        for col,value in zip(cols,expected):
            cell_value = ws[f"{col}{row}"].value
            if cell_value is None:
                cell_value = previous_header
            if not cell_value or value not in normalize(cell_value):
                errors.append(f"缺少或错误的表头: 期望 '{value}'，但找到了 '{cell_value}'")
            else:
                target_headers.append(value)
            previous_header = value

        # 如果有问题，抛出异常
        if errors:
            error_msg = "\n".join(errors)
            raise ValueError(f"Excel 校验失败，共发现 {len(errors)} 个问题：\n{error_msg}")

        logger.info("校验通过，未发现问题。")
        return target_headers

    async def load_data(self, file_path: str,**kwargs):
        """
        Load RISK data from a specified source.
        """
        db_client = kwargs.get("db_client")
        sop_id = kwargs.get("sop_id")
        tmp_dir = None
        try:
            tmp_dir, temp_file_path = await download_oss_to_temp(file_path)

            # 校验表头
            target_headers = self.validate_headers(file_contents="",file_path=temp_file_path)
            if db_client and sop_id:
                db_client.update_percent_by_id(sop_id, "30%", lang=kwargs.get("lang", "zh"))

            # 抽取结构化数据
            structured_rows = self.preprocess_data(file_path=temp_file_path, target_headers=target_headers)

            return structured_rows

        except Exception as e:
            logger.error(f"数据加载失败: {file_path}, 错误: {e}")
            raise
        finally:
            cleanup_tmp_dir(tmp_dir)

    def preprocess_data(self,file_path:str,**kwargs):
        """
        Preprocess the loaded RISK data.
        """
        # Implementation for preprocessing RISK data
        target_headers = kwargs.get("target_headers")
        wb = load_workbook(file_path)
        ws = wb.active
        cols = ['A','B', 'C', 'D', 'E','F','G']
        data_rows = []
        max_row = ws.max_row
        previous_row = {} # 上一轮的数据，用于填充空单元格
        for r in range(3, max_row + 1):
            row_dict = {}
            for col, header in zip(cols, target_headers):
                if row_dict.get(header,None):
                    row_dict[header] += "\n"+ws[f"{col}{r}"].value
                else:
                    row_dict[header] = ws[f"{col}{r}"].value

                    # 判断是否完全为空
            if all(v is None or str(v).strip() == "" for v in row_dict.values()):
                continue
            else:
                # 空单元格填充上一行
                for key, value in row_dict.items():
                    if value is None or str(value).strip() == "":
                        if key in previous_row:
                            row_dict[key] = previous_row[key]
            row_dict["行号"] = r
            content_list = []
            for header in target_headers:
                val = row_dict.get(header)
                val_str = str(val) if val is not None else ""
                content_list.append(f"{header}: {val_str}")
            row_dict["内容"] = " | ".join(content_list)
            previous_row = row_dict.copy()
            data_rows.append(row_dict.copy())
        return data_rows

class RISKPdfDataLoader(BaseDataLoader):
    """
    Data loader for RISK PDF documents.
    """

    def validate_headers(self,file_contents,**kwargs):
        """
        Abstract method to validate headers.
        """
        pass

    def load_data(self, file_path: str,**kwargs):
        """
        Load RISK data from a specified source.
        """
        pass

    def preprocess_data(self,file_path:str,**kwargs):
        """
        Preprocess the loaded RISK data.
        """
        pass


class OpGuideWordDataLoader(BaseDataLoader):
    """
    Data loader for Operation Guide Word documents.
    """

    async def validate_headers(self,file_contents,**kwargs):
        """
        Abstract method to validate headers.
        """
        pass

    async def load_data(self, file_path: str,**kwargs):
        """
        Load Operation Guide data from a specified source.
        """
        db_client = kwargs.get("db_client")
        sop_id = kwargs.get("sop_id")
        client = kwargs.get("client")
        if not client:
            client = get_client()
        tmp_dir = None
        temp_file_path = file_path
        try:
            text = await self.preprocess_data(temp_file_path)
            # 判断语种
            lang = await detect_language_from_preview(text[:1000], client)
            parser = DocHierarchyParser(lang=lang)
            if db_client and sop_id:
                db_client.update_percent_by_id(sop_id, "30%", lang=kwargs.get("lang", "zh"))
            hierarchy = await parser.aparse(
                text,
                client=client,
            )
            # 筛选
            nodes = []
            for section in hierarchy:
                title_content = section['title_content']
                # 排除包含 "." 或制表符的标题
                if not any(char in title_content for char in [".", "\t"]):
                    nodes.append(section)
            return nodes, lang
        except Exception as e:
            logger.error(f"Operation Guide Word 数据加载失败: {file_path}, 错误: {e}")
            raise

    async def preprocess_data(self,file_path:str,**kwargs):
        """
        Preprocess the loaded Operation Guide data.
        """
        if file_path.lower().endswith('.doc'):
            text, _ = _load_doc(file_path)
        elif file_path.lower().endswith('.docx'):
            text, _ = _load_docx(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {file_path}。仅支持 .doc 和 .docx 文件。")
        return text


class OpGuidePdfDataLoader(OpGuideWordDataLoader):
    """
    Data loader for Operation Guide PDF documents.
    """
    async def preprocess_data(self,file_path:str,**kwargs):
        """
        Preprocess the loaded Operation Guide data.
        """
        try:
            md_content = await self._parse_markdown_by_remote_magic_pdf(file_path)
        except Exception as remote_error:
            logger.warning(
                f"远端 magic-pdf 解析失败，回退本地解析: {file_path}, 错误: {remote_error}"
            )
            pdf_mid_data, md_content = magic_pdf_parse_util.pdf_parse_main(pdf_path=file_path)
        content = await self._remove_leading_hash_and_space(md_content)
        logger.info(f"content: {content}")
        return content

    async def _parse_markdown_by_remote_magic_pdf(self, file_path: str) -> str:
        return await asyncio.to_thread(self._parse_markdown_by_remote_magic_pdf_sync, file_path)

    def _parse_markdown_by_remote_magic_pdf_sync(self, file_path: str) -> str:
        if not self._is_remote_magic_pdf_enabled():
            raise RuntimeError("远端 magic-pdf 已禁用 (MAGIC_PDF_REMOTE_ENABLED=false)")

        parse_url = MAGIC_PDF_PARSE_URL
        parse_method = MAGIC_PDF_PARSE_METHOD
        parse_lang = MAGIC_PDF_PARSE_LANG
        parse_connect_timeout = MAGIC_PDF_PARSE_CONNECT_TIMEOUT
        parse_read_timeout = MAGIC_PDF_PARSE_READ_TIMEOUT
        parse_retries = MAGIC_PDF_PARSE_RETRIES
        retry_backoff = MAGIC_PDF_PARSE_RETRY_BACKOFF_SECONDS
        total_attempts = parse_retries + 1

        with open(file_path, "rb") as pdf_file:
            for attempt in range(1, total_attempts + 1):
                pdf_file.seek(0)
                try:
                    response = requests.post(
                        parse_url,
                        files={
                            "file": (
                                os.path.basename(file_path),
                                pdf_file,
                                "application/pdf",
                            )
                        },
                        data={"method": parse_method, "lang": parse_lang},
                        timeout=(parse_connect_timeout, parse_read_timeout),
                    )
                except requests.exceptions.RequestException as request_error:
                    if attempt < total_attempts and self._is_retryable_request_error(request_error):
                        logger.warning(
                            f"远端 magic-pdf 请求异常，准备重试: attempt={attempt}/{total_attempts}, url={parse_url}, error={request_error}"
                        )
                        time.sleep(retry_backoff * attempt)
                        continue
                    raise

                if response.status_code >= 400:
                    if attempt < total_attempts and self._is_retryable_http_status(response.status_code):
                        logger.warning(
                            f"远端 magic-pdf 返回可重试状态码，准备重试: attempt={attempt}/{total_attempts}, status={response.status_code}, url={parse_url}"
                        )
                        time.sleep(retry_backoff * attempt)
                        continue
                    response.raise_for_status()

                break

        try:
            payload: object = response.json()
        except ValueError:
            payload = response.text

        md_content = self._extract_markdown_content(payload)
        if not md_content:
            raise ValueError(f"远端 magic-pdf 响应缺少 Markdown 内容: {payload}")
        return md_content

    @staticmethod
    def _is_remote_magic_pdf_enabled() -> bool:
        return MAGIC_PDF_REMOTE_ENABLED

    @staticmethod
    def _is_retryable_http_status(status_code: int) -> bool:
        return status_code in {408, 429, 500, 502, 503, 504}

    @classmethod
    def _is_retryable_request_error(cls, error: Exception) -> bool:
        if isinstance(error, (requests.exceptions.Timeout, requests.exceptions.ConnectionError)):
            return True
        if isinstance(error, requests.exceptions.HTTPError):
            response = getattr(error, "response", None)
            if response is not None:
                return cls._is_retryable_http_status(response.status_code)
        return False

    @staticmethod
    def _extract_markdown_content(payload: object) -> str | None:
        candidate_keys = ("md_content", "markdown_content", "markdown", "content", "text")

        if isinstance(payload, str):
            return payload if payload.strip() else None

        if not isinstance(payload, dict):
            return None

        for key in candidate_keys:
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                return value

        data = payload.get("data")
        if isinstance(data, str) and data.strip():
            return data

        if isinstance(data, dict):
            for key in candidate_keys:
                value = data.get(key)
                if isinstance(value, str) and value.strip():
                    return value

        result = payload.get("result")
        if isinstance(result, dict):
            for key in candidate_keys:
                value = result.get(key)
                if isinstance(value, str) and value.strip():
                    return value
        return None

    async def _remove_leading_hash_and_space(self,text: str) -> str:
        """
        去除每一行最开头的 #，
        如果 # 后紧跟一个空格，也一并去掉。
        只处理真正的行首（不含前导空格）。
        """
        return re.sub(r'^# ?', '', text, flags=re.MULTILINE)


class ERDWordDataLoader(BaseDataLoader):
    """
    Data loader for ERD Word documents.
    """
    def validate_headers(self,file_contents,**kwargs):
        """
        Abstract method to validate headers.
        """
        # 判断是word文本内容还是都是表格内容（就读取出来判断里面能不能匹配到标题，并且数量大于3）
        table_loader = WordTableExtractorLite()
        hierarchy = table_loader.extract("", file_contents)
        if hierarchy and len(hierarchy) > 3:
            return True,hierarchy # 为表格内容格式
        else:
            return False,hierarchy # 为word内容格式

    async def load_data(self, file_path: str,**kwargs):
        """
        Load ERD data from a specified source.
        """
        db_client = kwargs.get("db_client")
        sop_id = kwargs.get("sop_id")
        client = kwargs.get("client")
        if not client:
            client = get_client()
        tmp_dir = None
        try:
            tmp_dir, temp_file_path = await download_oss_to_temp(file_path)
            if temp_file_path.lower().endswith('.doc'):
                text,doc = _load_doc(temp_file_path)
            elif temp_file_path.lower().endswith('.docx'):
                text,doc = _load_docx(temp_file_path)
            else:
                raise ValueError(f"不支持的文件格式: {file_path}。仅支持 .doc 和 .docx 文件。")
            lang = await detect_language_from_preview(text[:1000], client)
            flag,hierarchy = self.validate_headers(doc)
            if db_client and sop_id:
                db_client.update_percent_by_id(sop_id, "30%", lang=kwargs.get("lang", "zh"))
            if flag:
                return hierarchy, lang

            parser = DocHierarchyParser(lang=lang)
            parsed_sections = await parser.aparse(
                text,
                client=client,
            )
            return parsed_sections, lang
        except Exception as e:
            logger.error(f"ERD Word 数据加载失败: {file_path}, 错误: {e}")
            raise
        finally:
            cleanup_tmp_dir(tmp_dir)

    def preprocess_data(self,file_path:str,**kwargs):
        """
        Preprocess the loaded ERD data.
        """
        flag = kwargs.get("flag",True)
        hierarchy = kwargs.get("hierarchy","")
        text = kwargs.get("text","")
         # 根据不同的判断进行不同的数据转化
        if flag :
            return hierarchy
        # 需要根据word进行进一部的切分
        parser = DocHierarchyParser()
        hierarchy = parser.parse(text)
        return hierarchy


class ERDPdfDataLoader(BaseDataLoader):
    """
    Data loader for ERD PDF documents.
    """
    def validate_headers(self,file_contents,**kwargs):
        """
        Abstract method to validate headers.
        """
        # 判断是word文本内容还是都是表格内容（就读取出来判断里面能不能匹配到标题，并且数量大于3）
        parser = MagicPdfTableExtractor()
        tables = parser.extract(file_contents)
        if tables and len(tables) > 4:
            return True, tables  # 为表格内容格式
        else:
            return False, tables  # 为word内容格式

    async def load_data(self, file_path: str,**kwargs):
        """从 OSS 获取 PDF，判断内容结构后解析返回结果"""
        db_client = kwargs.get("db_client")
        sop_id = kwargs.get("sop_id")
        client = kwargs.get("client")
        if not client:
            client = get_client()
        tmp_dir = None
        temp_pdf_path = None
        try:
            tmp_dir, temp_pdf_path = await download_oss_to_temp(file_path, os.path.basename(file_path) or "input.pdf")
            pdf_mid_data, md_content = magic_pdf_parse_util.pdf_parse_main(pdf_path=temp_pdf_path)
            lang = await detect_language_from_preview(md_content[:1000], client)
            flag, tables = self.validate_headers(md_content)
            if db_client and sop_id:
                db_client.update_percent_by_id(sop_id, "30%", lang=kwargs.get("lang", "zh"))
            if flag:
                return tables, lang

            result = self.preprocess_data(file_path=temp_pdf_path, flag=flag, tables=tables, text=md_content)
            return result, lang
        except Exception as e:
            logger.error(f"ERD PDF 数据加载失败: {file_path}, 错误: {e}")
            raise
        finally:
            cleanup_tmp_dir(tmp_dir)

    def preprocess_data(self,file_path:str,**kwargs):
        """
        Preprocess the loaded ERD data.
        """
        flag = kwargs.get("flag",True)
        tables = kwargs.get("tables","")
        text = kwargs.get("text","")
         # 根据不同的判断进行不同的数据转化
        if flag :
            return tables
        # 需要根据表格进行进一部的切分
        word_loader = MarkdownHierarchyParser()
        return word_loader.parse(text)
