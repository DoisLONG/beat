# Copyright (C) 2026 Intel Corporation
# SPDX-License-Identifier: Apache-2.0

# Copyright (c) Opendatalab. All rights reserved.
import os
import json
from magic_pdf.data.data_reader_writer import FileBasedDataWriter, FileBasedDataReader
from magic_pdf.data.dataset import PymuDocDataset
from magic_pdf.model.doc_analyze_by_custom_model import doc_analyze
from magic_pdf.config.enums import SupportedPdfParseMethod
from loguru import logger
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# # args
# pdf_file_name = "/tmp/CAGRA.pdf"  # replace with the real pdf path
# name_without_suff = pdf_file_name.split(".")[0]

# # prepare env
# local_image_dir, local_md_dir = "output/images", "output"
# image_dir = str(os.path.basename(local_image_dir))

# os.makedirs(local_image_dir, exist_ok=True)

# image_writer, md_writer = FileBasedDataWriter(local_image_dir), FileBasedDataWriter(
#     local_md_dir
# )

# # read bytes
# reader1 = FileBasedDataReader("")
# pdf_bytes = reader1.read(pdf_file_name)  # read the pdf content

# # proc
# ## Create Dataset Instance
# ds = PymuDocDataset(pdf_bytes)

# ## inference
# if ds.classify() == SupportedPdfParseMethod.OCR:
#     infer_result = ds.apply(doc_analyze, ocr=True)

#     ## pipeline
#     pipe_result = infer_result.pipe_ocr_mode(image_writer)

# else:
#     infer_result = ds.apply(doc_analyze, ocr=False)

#     ## pipeline
#     pipe_result = infer_result.pipe_txt_mode(image_writer)



# ### draw model result on each page
# infer_result.draw_model(os.path.join(local_md_dir, f"{name_without_suff}_model.pdf"))

# ### get model inference result
# model_inference_result = infer_result.get_infer_res()

# ### draw layout result on each page
# pipe_result.draw_layout(os.path.join(local_md_dir, f"{name_without_suff}_layout.pdf"))

# ### draw spans result on each page
# pipe_result.draw_span(os.path.join(local_md_dir, f"{name_without_suff}_spans.pdf"))

# ### get markdown content
# md_content = pipe_result.get_markdown(image_dir)

# ### dump markdown
# pipe_result.dump_md(md_writer, f"{name_without_suff}.md", image_dir)

# ### get content list content
# content_list_content = pipe_result.get_content_list(image_dir)

# ### dump content list
# pipe_result.dump_content_list(md_writer, f"{name_without_suff}_content_list.json", image_dir)

# ### get middle json
# middle_json_content = pipe_result.get_middle_json()

# ### dump middle json
# pipe_result.dump_middle_json(md_writer, f'{name_without_suff}_middle.json')





def pdf_parse_main(
        pdf_path: str,
        parse_method: str = 'auto',
        output_dir: str = None,
):
    pdf_name = os.path.basename(pdf_path).split('.')[0]
    pdf_path_parent = os.path.dirname(pdf_path)
    if output_dir:
        output_path = os.path.join(output_dir, pdf_path)
    else:
        output_path = os.path.join(pdf_path_parent, pdf_name)

    output_image_path = os.path.join(output_path, 'images')
    # get the image_writer and md_writer
    image_writer, md_writer = FileBasedDataWriter(output_image_path), FileBasedDataWriter(output_path)
    image_dir = str(os.path.basename(output_image_path))

    # read bytes
    reader1 = FileBasedDataReader("")
    pdf_bytes = reader1.read(pdf_path)  # read the pdf content

    # proc
    ## Create Dataset Instance
    ds = PymuDocDataset(pdf_bytes)

    ## inference
    if parse_method == 'auto':
        if ds.classify() == SupportedPdfParseMethod.OCR:
            logger.info("Using OCR mode")
            infer_result = ds.apply(doc_analyze, ocr=True)
            pipe_result = infer_result.pipe_ocr_mode(image_writer)
        else:
            logger.info("Using TXT mode")
            infer_result = ds.apply(doc_analyze, ocr=False)
            pipe_result = infer_result.pipe_txt_mode(image_writer)
    elif parse_method == 'ocr':
        logger.info("Using OCR mode")
        infer_result = ds.apply(doc_analyze, ocr=True)
        pipe_result = infer_result.pipe_ocr_mode(image_writer)
    elif parse_method == 'txt':
        logger.info("Using TXT mode")
        infer_result = ds.apply(doc_analyze, ocr=False)
        pipe_result = infer_result.pipe_txt_mode(image_writer)
    else:
        raise ValueError(f"Unsupported parse method: {parse_method}")

    ### get middle json
    middle_json_content = pipe_result.get_middle_json()
    ### get markdown content
    md_content = pipe_result.get_markdown(image_dir)
    return middle_json_content,md_content


def extract_page_info(data):
    # Add error handling for None input
    if data is None:
        logger.error("Input data is None")
        return []

    result = []

    # Get pdf_info with default empty list if not found
    json_data = json.loads(data)
    pdf_info = json_data["pdf_info"]
    if not pdf_info:
        logger.warning("No pdf_info found in input data")
        return []

    for page_idx, page in enumerate(pdf_info):
        zones = []
        blocks = []
        # Extracting para_blocks
        para_blocks = page.get("para_blocks", [])
        if para_blocks:
            for block in para_blocks:
                blocks_out = block.get("blocks", [])
                if blocks_out:
                    blocks.extend(blocks_out)
                else:
                    blocks.append(block)
        i=0;
        for block in blocks:
            zone = {
                "rect": block.get("bbox", []),
                "type": block.get("type", ""),
                "is_discard": False,
                "lines": [],
                "index": i,
            }
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    zone["lines"].append({
                        "text": span.get("content", ""),
                        "rect": span.get("bbox", []),
                        "type": span.get("type", ""),
                        "is_discard": False
                    })
            zones.append(zone)
            i +=1
        # Extracting discarded_blocks
        for block in page.get("discarded_blocks", []):
            zone = {
                "rect": block.get("bbox", []),
                "type": block.get("type", ""),
                "is_discard": True,
                "lines": []
            }

            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    zone["lines"].append({
                        "text": span.get("content", ""),
                        "rect": span.get("bbox", []),
                        "type": span.get("type", ""),
                        "is_discard": True
                    })
            zones.append(zone)

        # # Sort zones by y1 (top-left corner) and then by x1 (left)
        # valid_zones = [zone for zone in zones if 'rect' in zone and len(zone['rect']) >= 2]
        # valid_zones.sort(key=lambda x: (x['rect'][1], x['rect'][0]))
        # # zones.sort(key=lambda x: (x['rect'][1], x['rect'][0]))
        # zones = valid_zones

        # Page info structure
        page_info = {
            "page_info": {
                "page_no": page_idx + 1,
                "page_size": {
                    "width": page["page_size"][0],
                    "height": page["page_size"][1]
                }
            },
            "zones": zones
        }

        result.append(page_info)

    return result

def pdf_table_to_excel(pdf_path, excel_path, min_col_match=0.8):
    """
    兼容超旧版本pdfplumber（无bbox参数）：通过列数和表头判断跨页表格
    """
    all_tables = []  # 存储所有表格数据：[表格1数据, 表格2数据, ...]
    prev_table = None  # 上一页的表格数据（用于跨页判断）

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            # 提取当前页所有表格（旧版本直接用extract_tables()）
            current_tables = page.extract_tables() or []  # 空列表处理
            if not current_tables:
                prev_table = None
                continue

            for table in current_tables:
                if not table:  # 跳过空表格
                    continue

                # 步骤1：判断是否为跨页表格（通过列数和表头相似度）
                if prev_table is not None:
                    # 条件1：列数需基本一致（允许±1差异）
                    prev_cols = len(prev_table[0]) if prev_table else 0
                    curr_cols = len(table[0]) if table else 0
                    col_diff = abs(prev_cols - curr_cols) <= 1

                    # 条件2：表头（第一行）相似度超过阈值
                    if prev_table and table and prev_cols > 0 and curr_cols > 0:
                        # 取表头最小长度（避免索引越界）
                        min_head_len = min(prev_cols, curr_cols)
                        # 计算相同元素数量
                        same_head = sum(
                            str(prev_table[0][i]).strip() == str(table[0][i]).strip()
                            for i in range(min_head_len)
                            if prev_table[0][i] is not None and table[0][i] is not None
                        )
                        head_similarity = same_head / min_head_len if min_head_len > 0 else 0
                    else:
                        head_similarity = 0

                    # 满足列数和表头条件，判定为跨页表格
                    if col_diff and head_similarity >= min_col_match:
                        all_tables.pop()  # 移除上一页表格
                        merged_table = prev_table + table  # 合并数据
                        all_tables.append(merged_table)
                        prev_table = merged_table
                        continue

                # 非跨页表格：直接添加
                all_tables.append(table)
                prev_table = table

    # 步骤2：写入Excel基础数据
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for i, table_data in enumerate(all_tables, start=1):
            df = pd.DataFrame(table_data)
            sheet_name = f"table_{i}"
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

    # 步骤3：还原合并单元格和调整格式
    wb = load_workbook(excel_path)
    for i, table_data in enumerate(all_tables, start=1):
        sheet = wb[f"table_{i}"]
        rows = len(table_data)
        cols = len(table_data[0]) if rows > 0 else 0

        # 还原合并单元格（同之前逻辑）
        merge_ranges = []
        for col in range(cols):
            start_row = None
            for row in range(rows):
                cell_val = table_data[row][col]
                if cell_val is not None and start_row is None:
                    start_row = row
                elif cell_val is None and start_row is not None:
                    end_row = row
                    while end_row + 1 < rows and table_data[end_row + 1][col] is None:
                        end_row += 1
                    if end_row > start_row:
                        merge_range = f"{get_column_letter(col+1)}{start_row+1}:{get_column_letter(col+1)}{end_row+1}"
                        merge_ranges.append(merge_range)
                    start_row = None
                elif cell_val is not None and start_row is not None:
                    start_row = row

        for merge_range in merge_ranges:
            sheet.merge_cells(merge_range)

        # 调整列宽
        for col in range(cols):
            if rows == 0:
                max_len = 10
            else:
                max_len = max(
                    len(str(sheet.cell(row=row+1, column=col+1).value) or "")
                    for row in range(rows)
                )
            sheet.column_dimensions[get_column_letter(col+1)].width = min(max_len + 2, 50)

    wb.save(excel_path)
    logger.info(f"转换完成！共提取 {len(all_tables)} 个表格，保存至：{excel_path}")

# 测试
if __name__ == '__main__':
    current_script_dir = os.path.dirname(os.path.abspath(__file__))
    # demo_names = ['demo1', 'demo2', 'small_ocr']
    md_content = None
    # demo_names = ['/tmp/CAGRA_6','/tmp/中国启蒙运动A-10']
    # demo_names = ['/tmp/中国启蒙运动A-10']
    demo_names = ['/opt/cn/原料工区操作规程（配料集操工）.pdf']

    # 调用示例
    pdf_table_to_excel("/opt/cn/原料工区操作规程（配料集操工）.pdf", "/opt/cn/output.xlsx")

    # for name in demo_names:
    #     file_path = os.path.join(current_script_dir, f'{name}.pdf')
    #     # page_bytes = extract_pdf_pages_as_bytes(file_path,-1,-1)
    #     # logger.debug(f"length page_bytes: {len(page_bytes)}")
    #     # pdf_mid_data,md_content = pdf_parse_main(file_path,beginPageNo=0,endPageNo=-1)
    #     pdf_mid_data,md_content = pdf_parse_main(file_path,parse_method="ocr")
    #     logger.debug(f"pdf_mid_data: {pdf_mid_data}")
    #     logger.debug(f"md_content: {md_content}")
    #     # pdf_mid_data,md_content = pdf_parse_main(file_path)
    #
    #     # logger.debug(f"pdf_mid_data: {pdf_mid_data}")
    #
    #     pages_info = extract_page_info(pdf_mid_data)
    #     pretty_json = json.dumps(pages_info,ensure_ascii=False, indent=4)
    #     logger.debug(f"pdf_mid_data: {json.dumps(pdf_mid_data,ensure_ascii=False, indent=4)}")
    #     logger.debug(f"pages_info: {pretty_json}")
    #     logger.debug(f"md_content: {md_content}")