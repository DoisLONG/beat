# Copyright (C) 2026 Intel Corporation
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def create_excel_from_structure(
    table_data: dict[str, Any],
    output_path: Path,
    title: str = "",
) -> Path:
    workbook = Workbook()
    worksheet = workbook.active

    if title:
        worksheet.title = re.sub(r'[\\/:*?"<>|]', "", title[:31])

    headers = table_data.get("headers", ["序号", "内容"])
    rows = table_data.get("rows", [])

    header_start_row = 1
    if title:
        worksheet.cell(row=1, column=1, value=title)
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)
        if len(headers) > 1:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        header_start_row = 2

    for column_index, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_start_row, column=column_index, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(fill_type="solid", start_color="366092")

    for row_index, row_data in enumerate(rows, header_start_row + 1):
        values = row_data if isinstance(row_data, list) else [str(row_data)] if row_data else [""]
        for column_index, value in enumerate(values, 1):
            if column_index > len(headers):
                break
            cell = worksheet.cell(row=row_index, column=column_index, value=str(value))
            if row_index % 2 == 0:
                cell.fill = PatternFill(fill_type="solid", start_color="F2F2F2")

    for column_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(column_index)
        max_length = 0
        for row in worksheet.iter_rows(
            min_row=header_start_row,
            max_row=worksheet.max_row,
            min_col=column_index,
            max_col=column_index,
        ):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row in worksheet.iter_rows(
        min_row=header_start_row,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=len(headers),
    ):
        for cell in row:
            cell.border = thin_border

    worksheet.freeze_panes = worksheet.cell(row=header_start_row + 1, column=1)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path
